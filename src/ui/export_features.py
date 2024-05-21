import shutil
import os
from PyQt5.QtWidgets import QFileDialog
from openpyxl import load_workbook

class ExportFeatures:
    """
    This class handles the functionality of exporting coordinates
    from the UI Table to xslx format using the provided template. 
    """
    
    def __init__(self, ui):
        self.ui = ui
        self.setup_connections()

    def setup_connections(self):
        self.ui.actionExport_to_xls.triggered.connect(self.handle_export_action_triggered)

    def handle_export_action_triggered(self):    
        # Select file name and location
        self.export_dialog()
        
    def export_dialog(self):
        """
        Method which opens a file dialog to allow user to select ouput
        location & name for kinematics data export.
        Copies the 'template' Excel file and places it in specified location,
        in order to be updated by user. 
        """
        # Dynamically get the directory of the current script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        # Navigate to root directory of project
        project_root_dir = os.path.abspath(os.path.join(script_dir, "..", ".."))
        # Initial directory for FileDialog (output folder)
        initial_dir = os.path.join(project_root_dir, "output")
        
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getSaveFileName(self.ui.centralwidget, "Save XLSX File", initial_dir, "Excel Files (*.xlsx)", options=options)
        if file_name:

            # Construct the full path to the template file
            template_path = os.path.join(project_root_dir, "input", "Excel_Input_Template.xlsx")

            if os.path.exists(template_path):
                # Copy the template file to the specified location
                shutil.copy(template_path, file_name)                
                self.export_file_name = file_name
                self.export_to_excel()
            else:
                print(f"Template Excel file not found at: {template_path}. Export not possible.")         
                
    def export_to_excel(self):
        """
        Retreives kinematic data from UITable (both front and rear)
        and updates copied template Excel file (from export_dialog())
        """        
        # Open workbook   
        export_workbook = load_workbook(self.export_file_name)
        sheet = export_workbook.active
        
        # 'Front' tab
        for xlsx_row, table_row in zip(range(3, 9), range(0, 6)): # Iterate over rows
            for xlsx_col, table_col in zip(range(2, 5), range(0, 3)): # Iterate over columns 
                # Retrieve value from UITable
                item = self.ui.frontInput.item(table_row, table_col)
                if item:  # Check if the value exists in UITable
                    value = float(item.text())  # Get the text of the item
                    # Add to xlsx
                    sheet.cell(row=xlsx_row, column=xlsx_col, value=value)
                
                else:
                    print(f"No coordinate found for Front {self.ui.frontInput.item(table_row, 0)}")
                        
        # 'Rear' tab
        for xlsx_row, table_row in zip(range(3, 9), range(0, 6)): # Iterate over rows
            for xlsx_col, table_col in zip(range(5, 8), range(0, 3)): # Iterate over columns 
                # Retrieve value from UITable
                item = self.ui.rearInput.item(table_row, table_col)
                if item:  # Check if the value exists in UITable
                    value = float(item.text())  # Get the text of the item
                    # Add to xlsx
                    sheet.cell(row=xlsx_row, column=xlsx_col, value=value)
                else:
                    print(f"No coordinate found for Rear {self.ui.rearInput.item(table_row, 0)}")
                
        # Save the workbook
        export_workbook.save(self.export_file_name)
        print(f"Kinematic data exported to {self.export_file_name}")
        
        
        
        
        
        
        
                
                
    