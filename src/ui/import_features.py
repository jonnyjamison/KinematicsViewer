from PyQt5.QtWidgets import QFileDialog
from openpyxl import load_workbook
from PyQt5.QtWidgets import QTableWidgetItem


class ImportFeatures:
    """
    This class handles the functionality of importing coordinates
    from the provided xlsx template to the UI Table. 
    """
    
    def __init__(self, ui):
        self.ui = ui 
        self.setup_connections()

    def setup_connections(self):
        self.ui.actionImport_from_xls.triggered.connect(self.handle_import_action_triggered)

    def handle_import_action_triggered(self):
        """
        Opens a file dialog which allows user to select excel file
        to populate UI Table kinematics data.
        """
        # Select xlsx file for import using file dialouge
        file_dialog = QFileDialog(self.ui.centralwidget)
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        file_dialog.setNameFilter("Excel Files (*.xlsx)")
        file_dialog.setOption(QFileDialog.DontUseNativeDialog, True)  # Ensure single file selection
        if file_dialog.exec_():
            selected_file = file_dialog.selectedFiles()
            if selected_file:
                file_path = selected_file[0]
                print("Selected input file:", file_path)
                self.read_excel_data(file_path)

    def read_excel_data(self, file_path):
        """
        Iterates through front and rear sections of Excel
        and assigns data to the 'front' and 'rear' tabs of the
        UI Table respectively. 
        """
        workbook = load_workbook(filename=file_path)
        # Read data from the first sheet
        xlsx_data = workbook.active
                                
        # Add data to kinData UI Table
        # For 'Front' tab 
        for xlsx_row, table_row in zip(range(3, 9), range(0, 6)): # Iterate over rows
            for xlsx_col, table_col in zip(range(2, 5), range(0, 3)): # Iterate over columns 
                # Retrieve the value from the xlsx file
                value = xlsx_data.cell(row=xlsx_row, column=xlsx_col).value
                # Create a QTableWidgetItem with the values
                item = QTableWidgetItem(str(value))
                # Set the item in the table at the corresponding row and column
                self.ui.frontInput.setItem(table_row, table_col, item)
                
        # For 'Rear' tab  
        for xlsx_row, table_row in zip(range(3, 9), range(0, 6)): # Iterate over rows
            for xlsx_col, table_col in zip(range(5, 8), range(0, 3)): # Iterate over columns
                # Retrieve the value from the xlsx file
                value = xlsx_data.cell(row=xlsx_row, column=xlsx_col).value
                # Create a QTableWidgetItem with the value
                item = QTableWidgetItem(str(value))
                # Set the item in the table at the corresponding row and column
                self.ui.rearInput.setItem(table_row, table_col, item)